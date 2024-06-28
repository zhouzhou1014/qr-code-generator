#!/usr/bin/env python
# -*- coding: UTF-8 -*-
# @Time    : 2024/6/21 21:23
# @Author  : zhouzhou
# @File    : my_execl.py
# @Software: PyCharm
import os
from openpyxl import load_workbook


# from my_path import *
class MyExcel:
    # excel_path 文件路径
    # sheet_name sheet名字
    # def __init__(self, excel_path, sheet_name):
    #     # 1、加载一个excel，得到工作薄 Workbook
    #     wb = load_workbook(excel_path)
    #     # 2、选择一个表单- 通过表单名 Sheet
    #     self.sh = wb[sheet_name]

    # 函数读取Excel文件并返回工作表名称列表
    def get_sheet_names(self, file_path):
        workbook = load_workbook(file_path)
        return workbook.sheetnames

    def read_data(self, excel_path, sheet_name):
        # 注意：接口的请求数据，读取出来是字符串。
        # 存储表单下读取到的所有数据 - 每一个成员都是一个字段
        # 1、加载一个excel，得到工作薄 Workbook
        wb = load_workbook(excel_path)
        # 2、选择一个表单- 通过表单名 Sheet
        self.sh = wb[sheet_name]
        all_data = []
        data = list(self.sh.values)
        keys = data[0]  # 获取所有的列名
        # data[1:] 从文件第二行开始读取数据,获取用例
        for row in data[1:]:
            # 使用dict函数转成字典格式，使用zip函数将列名和行数据一一对应
            row_dict = dict(zip(keys, row))
            # 使用append函数写入all_data列表中
            all_data.append(row_dict)
            # 使用return返回值
        return all_data

    def write_to_excel(self, data_path, sheet_name, header_row, data_rows):
        """
        :param data_path: 文件路径
        :param sheet_name: sheet名字
        :param header_row: 列名行
        :param data_rows: 数据行
        创建 Excel 表格和写入数据。
        在创建 Excel 表格时，
        使用了 openpyxl 库中的 create_sheet() 方法创建了一个新的工作表，
        并使用 cell() 方法将表头写入到指定的单元格中。在写入数据时，
        使用了两个嵌套的循环来遍历数据行和列，
        并使用 cell() 方法将数据写入到指定的单元格中。
        最后，使用 save() 方法将 Excel 文件保存到指定的路径中。
        sheet_name = f"my_Video_homework"
        gauge_outfit_list = ['videoId', 'Title', 'courseId']
        rows = [('1219199782', '1-1 The Etiquette of Meeting Strangers 与陌生人见面的礼节', '5956249')]
        """
        workbook = load_workbook(filename=data_path)
        sheet0 = workbook.create_sheet(sheet_name, 1)
        for col, header in enumerate(header_row, start=1):
            sheet0.cell(row=1, column=col, value=header)
        for row, data in enumerate(data_rows, start=2):
            for col, value in enumerate(data, start=1):
                sheet0.cell(row=row, column=col, value=value)
        workbook.save(data_path)
        return '写入成功'


if __name__ == '__main__':
    # openxls_create()
    # data_path = os.path.join(BasicData_path, 'shop_02.xlsx')
    #
    # basedir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    # me = MyExcel().read_data(data_path, "76658448")
    # # cases = me.read_data()
    # print(me)
    # rows = []
    # a = [{'id': '76933790'},{'id': '76933790'}]
    # for i in me:
    #     id = i['id']
    #     rows.extend(id)
    # print(rows)
    a = [{'id': '76933790'}, {'id': '76933790'}]
